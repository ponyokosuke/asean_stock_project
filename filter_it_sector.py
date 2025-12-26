import pandas as pd
import yfinance as yf
import os
import time
import json
import google.generativeai as genai
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# --- 設定 ---
# 1. APIキーの読み込み
load_dotenv()
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
if GEMINI_API_KEY:
    genai.configure(api_key=GEMINI_API_KEY)
else:
    print("Error: GEMINI_API_KEY is not set in the .env file.")
    exit()

# 2. 銘柄リストのインポート
try:
    import asean_stock_codes
    ALL_CODES = asean_stock_codes.ALL_ASEAN_CODES
except ImportError:
    print("Warning: asean_stock_codes.py not found. Using a test list.")
    ALL_CODES = ["D05.SI", "Z74.SI", "4863.KL", "0021.KL"] 

# 3. 高精度プロンプト (完全英語)
IT_JUDGEMENT_PROMPT = """
You are a financial analyst specializing in technology sector classification.
Your task is to analyze the provided 'Summary of Business' for multiple companies and determine if they qualify as an **"IT-related Company"** based on the strict criteria below.

### 1. CRITERIA FOR INCLUSION (Verdict: "Yes")
Classify as "Yes" ONLY if the company's CORE business falls into one of these categories:
* **Software & IT Services:** Software development, SaaS, ERP, CRM, System Integration (SI), IT Consulting, Managed Services, Cybersecurity, AI, IoT.
* **Hardware & Technology Equipment:** Manufacturing or distributing enterprise IT hardware (servers, network equipment), Semiconductors, Electronic components for computing.
* **Telecommunications:** Telecommunication carriers, ISPs, Network infrastructure providers, Data Centers.

### 2. CRITERIA FOR EXCLUSION (Verdict: "No")
Classify as "No" if the company is merely a USER of IT, or if IT is secondary:
* **Fintech / Digital Banking:** (Exclude if core is financial services).
* **E-commerce Retail:** (Exclude if primarily selling physical goods).
* **General Manufacturing:** (Exclude unless strictly IT-related).

### 3. CRITERIA FOR AMBIGUITY (Verdict: "Grey")
Classify as "Grey" if the company has a mix of IT and non-IT businesses, and it is difficult to determine which is dominant, or if the summary is too vague.

### 4. OUTPUT FORMAT
Return a **valid JSON object**. 
The keys must be the STOCK_CODE.
The values must be an object with: "verdict", "category", and "reason".

{
  "STOCK_CODE": {
    "verdict": "Yes",  // Options: "Yes", "No", "Grey"
    "category": "Software", // Options: "Software", "IT Services", "Hardware", "Telecom", "Mixed", "N/A"
    "reason": "A concise explanation in **ENGLISH** (1-2 sentences) justifying the verdict." 
  }
}
"""

def fetch_summaries(codes):
    """Yahoo FinanceからSummaryを取得する"""
    data_list = []
    print(f"Total target stocks: {len(codes)}")
    
    for i, code in enumerate(codes):
        print(f"\rFetching data: {i+1}/{len(codes)} ({code})", end="")
        try:
            ticker = yf.Ticker(code)
            info = ticker.info
            summary = info.get('longBusinessSummary', '')
            name = info.get('longName', code)
            
            if summary:
                data_list.append({
                    "code": code,
                    "name": name,
                    "summary": summary
                })
        except Exception:
            pass
    print("\nData fetch complete.")
    return data_list

def batch_judge_it_sector(targets):
    """LLMにまとめて投げて判定させる（リトライ機能付き）"""
    
    all_results = [] # Yes/No/Grey すべて格納するリスト
    batch_size = 50  # まとめて送る数
    
    model = genai.GenerativeModel('gemini-2.5-flash') 
    
    print(f"\nStarting AI Analysis: Analyzing {len(targets)} companies...")
    
    for i in range(0, len(targets), batch_size):
        batch = targets[i : i + batch_size]
        print(f"  - Analyzing batch: {i+1} to {min(i+batch_size, len(targets))}...")
        
        # 入力データの作成
        input_data_text = ""
        for item in batch:
            s = item['summary'][:800].replace("\n", " ")
            input_data_text += f"Code: {item['code']}\nSummary: {s}...\n---\n"
            
        full_prompt = f"""
        {IT_JUDGEMENT_PROMPT}

        ### TARGET COMPANIES DATA
        {input_data_text}
        """
        
        # リトライロジック
        max_retries = 3
        for attempt in range(max_retries):
            try:
                # AI実行
                response = model.generate_content(full_prompt)
                clean_text = response.text.replace("```json", "").replace("```", "").strip()
                result_json = json.loads(clean_text)
                
                # 成功したらデータを格納
                for item in batch:
                    code = item['code']
                    if code in result_json:
                        res = result_json[code]
                        verdict = res.get("verdict", "No")
                        category = res.get("category", "N/A")
                        
                        # ★変更点: Yesだけでなく、全ての結果をリストに追加する
                        all_results.append({
                            "Code": code,
                            "Name": item['name'],
                            "Verdict": verdict,
                            "Category": category,
                            "Reason": res.get("reason")
                        })
                        
                        # ログ出力（Yesのときだけ目立たせる）
                        if verdict == "Yes":
                            print(f"    [HIT] {code}: {category}")
                        # else:
                        #     print(f"    [SKIP] {code}: {verdict}") # 必要ならコメントアウト解除
                
                time.sleep(2) 
                break 

            except Exception as e:
                error_msg = str(e)
                if "429" in error_msg or "quota" in error_msg.lower():
                    wait_time = 30 
                    print(f"    ⚠️ Rate limit exceeded. Waiting {wait_time} seconds before retrying... (Attempt {attempt+1}/{max_retries})")
                    time.sleep(wait_time)
                else:
                    print(f"    ⚠️ Error in batch: {e}")
                    break 
            
    return all_results

def save_to_excel(data_list, filename):
    """詳細データをExcel形式で保存する（Yes/No/Grey全件）"""
    
    # データを判定順に並べ替え (Yes -> Grey -> No)
    sorter = {"Yes": 0, "Grey": 1, "No": 2}
    data_list.sort(key=lambda x: sorter.get(x["Verdict"], 3))
    
    df = pd.DataFrame(data_list)
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='All_Judgements')
        
        workbook = writer.book
        worksheet = writer.sheets['All_Judgements']
        
        # ヘッダーの装飾
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            
        # 列幅の設定
        column_widths = {'A': 10, 'B': 35, 'C': 10, 'D': 20, 'E': 70}
        for col_char, width in column_widths.items():
            worksheet.column_dimensions[col_char].width = width
            
        # セルのスタイル設定
        center_align = Alignment(horizontal='center', vertical='center')
        wrap_align = Alignment(wrap_text=True, vertical='center')
        
        for row in worksheet.iter_rows(min_row=2, max_row=len(data_list)+1):
            # Verdict (C列) の色分け
            verdict_cell = row[2] 
            verdict_cell.alignment = center_align
            val = verdict_cell.value
            
            if val == "Yes":
                verdict_cell.font = Font(color="006100", bold=True) # 緑
            elif val == "No":
                verdict_cell.font = Font(color="9C0006") # 赤
            elif val == "Grey":
                verdict_cell.font = Font(color="9C5700", bold=True) # オレンジ

            # Reason (E列) は折り返し
            row[4].alignment = wrap_align

    print(f"Excel file saved: {filename}")

def save_code_only_csv(data_list, filename):
    """main.py連携用：Yesの証券コードだけをCSV保存する"""
    # Yesだけのリストを作成
    yes_codes = [item['Code'] for item in data_list if item['Verdict'] == "Yes"]
    
    df = pd.DataFrame(yes_codes)
    df.to_csv(filename, index=False, header=False, encoding='utf-8')
    print(f"CSV file (YES codes only) saved: {filename}")
    print(f"Number of target companies for main.py: {len(yes_codes)}")

def main():
    # 1. 対象国の選択
    target_country = input("Enter Target Country Code (e.g., MY, SG, or ALL): ").strip().upper()
    
    target_codes = []
    suffix_map = {
        "MY": ".KL", "SG": ".SI", "ID": ".JK", "TH": ".BK", "VN": ".VN", "PH": ".PS"
    }
    
    if target_country == "ALL":
        target_codes = ALL_CODES
    elif target_country in suffix_map:
        suffix = suffix_map[target_country]
        target_codes = [c for c in ALL_CODES if c.endswith(suffix)]
    else:
        print("Searching all codes or specific suffix provided manually...")
        target_codes = ALL_CODES

    # 2. Summary取得
    print("Retrieving Business Summaries from Yahoo Finance...")
    data_with_summary = fetch_summaries(target_codes)
    
    if not data_with_summary:
        print("No business summaries retrieved.")
        return

    # 3. AI判定 (全件取得)
    all_results = batch_judge_it_sector(data_with_summary)
    
    # 4. ファイル保存
    if all_results:
        date_str = pd.Timestamp.now().strftime('%Y%m%d')
        
        # A. 詳細Excel (Yes/No/Grey 全件)
        excel_filename = f"IT_Judgement_Report_{target_country}_{date_str}.xlsx"
        save_to_excel(all_results, excel_filename)
        
        # B. コードのみCSV (Yesのみ、財務データ取得用)
        csv_filename = f"IT_Targets_{target_country}_{date_str}.csv"
        save_code_only_csv(all_results, csv_filename)
        
        print("\n------------------------------------------------")
        print(f"Analysis Complete.")
        print(f"Total Companies Analyzed: {len(all_results)}")
        print(f"1. Full Report (Excel): {excel_filename}")
        print(f"2. Target List (CSV):   {csv_filename}")
        print("------------------------------------------------")
        
        print(f"\n[Next Step] To download financial data for 'Yes' companies, run:")
        print(f"python main.py {csv_filename}")
        
    else:
        print("No results returned from AI.")

if __name__ == "__main__":
    main()