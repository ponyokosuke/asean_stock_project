import pandas as pd
import datetime
from pathlib import Path
import sys
import time
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font

import data_processor

def main():
    if len(sys.argv) < 2:
        print("------------------------------------------------")
        print("エラー: 読み込むCSVファイル名を指定してください。")
        print("実行例: python main.py asean_list.csv")
        print("------------------------------------------------")
        return

    csv_file_to_load = sys.argv[1]
    
    try:
        df_input = pd.read_csv(csv_file_to_load, header=None)
        codes = df_input[0].astype(str).tolist()
        print(f"📂 {csv_file_to_load} を読み込みました。対象: {len(codes)} 銘柄")
    except Exception as e:
        print(f"❌ CSV読み込みエラー: {e}")
        return

    print("\n=== ASEAN株 財務データ取得システム (AIセグメント分析付き) ===")
    
    all_results = []

    for code in codes:
        code = code.strip()
        print(f"\n--- {code} の処理中 ---")
        
        raw_data = data_processor.get_stock_data(code)
        
        if raw_data:
            processed_data = data_processor.extract_data(code, raw_data)
            all_results.append(processed_data)
            print(f"  会社名: {processed_data.get('Name of Company')}")
            print(f"  売上高: {processed_data.get('REVENUE')}")
        else:
            print("  データの取得に失敗しました。")
        
        time.sleep(0.5)

    if all_results:
        print("\n--- 全データ取得完了。AIによるセグメント分析を開始します ---")
        all_results = data_processor.batch_analyze_segments(all_results)

    if all_results:
        print("\nExcelファイルを作成しています...")
        df = pd.DataFrame(all_results)
        
        # 重複列の削除
        df = df.loc[:, ~df.columns.duplicated()]
        df = data_processor.format_for_excel(df)
        
        if "Sector /Industry" in df.columns:
            df = df.drop(columns=["Sector /Industry"])
            
        df["Ref"] = range(1, len(df) + 1)

        empty_cols = [
            "Taka's comments", "Remarks", "Visited (V) / Meeting Proposal (MP)",
            "Access", "Last Communications", "Category Classification/\nShareInvestor", 
            "Incorporated\n (IN / Year)", "Category Classification/SGX", "Sector & Industry/ SGX"
        ]
        for col in empty_cols:
            df[col] = ""
        
        df["Listed 'o' / Non Listed \"x\""] = "o"

        # ★変更: 前日の日付を計算して表示する
        # 現在日時から1日引く
        yesterday = datetime.datetime.now() - datetime.timedelta(days=1)
        yesterday_str = yesterday.strftime("%b %d") # 例: Dec 28
        
        # 株価カラム名: (Dec 28, Closing)
        final_stock_price_col = f"Stock Price ({yesterday_str}, Closing)"
        if "Stock Price" in df.columns:
            df = df.rename(columns={"Stock Price": final_stock_price_col})
            
        # 為替レートカラム名: (Dec 28, Closing)
        final_rate_col = f"Exchange Rate (to SGD) ({yesterday_str}, Closing)"
        if "Exchange Rate" in df.columns:
            df = df.rename(columns={"Exchange Rate": final_rate_col})

        target_order = [
            "Ref", "Name of Company", "Code", "Listed 'o' / Non Listed \"x\"",
            "Taka's comments", "Remarks", "Visited (V) / Meeting Proposal (MP)",
            "Website", "Major Shareholders", "Currency", 
            final_rate_col, # ★為替レート
            "FY", "REVENUE SGD('000)", "Segments", "PROFIT ('000)",
            "GROSS PROFIT ('000)", "OPERATING PROFIT ('000)",
            "NET PROFIT (Group) ('000)", "NET PROFIT (Shareholders) ('000)",
            "Minority Interest ('000)", "Shareholders' Equity ('000)",
            "Total Equity ('000)", "TOTAL ASSET ('000)", "Debt/Equity(%)",
            "Loan ('000)", "Loan/Equity (%)",
            final_stock_price_col, # ★株価
            "Shares Outstanding ('000)", "Market Cap ('000)",
            "Summary of Business", "Chairman / CEO", "Address", "Contact No.",
            "Access", "Last Communications", "Number of Employee Current",
            "Category Classification/YahooFin", "Sector & Industry/YahooFin",
            "Category Classification/\nShareInvestor", "Incorporated\n (IN / Year)",
            "Category Classification/SGX", "Sector & Industry/ SGX"
        ]
        
        for col in target_order:
             if col not in df.columns:
                 df[col] = ""
        
        if "Number of Employee" in df.columns:
            df = df.rename(columns={"Number of Employee": "Number of Employee Current"})
        
        df = df.loc[:, ~df.columns.duplicated()]
        df = df.reindex(columns=target_order)

        today = datetime.date.today().strftime("%Y-%m-%d")
        base_name = f"asean_financial_data_{today}"
        filename = f"{base_name}.xlsx"
        
        counter = 1
        while Path(filename).exists():
            filename = f"{base_name}_{counter}.xlsx"
            counter += 1
            
        try:
            df.to_excel(filename, index=False)
            
            wb = load_workbook(filename)
            ws = wb.active
            right_align = Alignment(horizontal='right')
            
            # 背景色: #fefe99
            header_fill = PatternFill(start_color="fefe99", end_color="fefe99", fill_type="solid")
            header_font = Font(bold=True)

            for cell in ws[1]:
                col_name = str(cell.value)
                col_idx = cell.column
                cell.fill = header_fill
                cell.font = header_font
                
                number_format = None
                apply_alignment = False
                
                if "('000)" in col_name:
                    number_format = '#,##0;(#,##0)'
                    apply_alignment = True
                elif "(%)" in col_name or "%" in col_name:
                    number_format = '0.00%'
                    apply_alignment = True
                elif col_name == "FY":
                    apply_alignment = True
                elif "Stock Price" in col_name:
                    number_format = '#,##0.000'
                    apply_alignment = True
                elif "Exchange Rate" in col_name:
                    number_format = '0.0000'
                    apply_alignment = True
                
                if number_format or apply_alignment:
                    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                        for cell_data in row:
                            if apply_alignment:
                                cell_data.alignment = right_align
                            if number_format:
                                cell_data.number_format = number_format
            
            wb.save(filename)
            print(f"★★★ 成功: {filename} に保存しました ★★★")
            
        except Exception as e:
            print(f"エラー: Excel保存に失敗しました ({e})")
    else:
        print("保存するデータがありませんでした。")

if __name__ == "__main__":
    main()