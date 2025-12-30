import streamlit as st
import pandas as pd
import time
import io
import os
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from dotenv import load_dotenv # 追加

# .envファイルを読み込む
load_dotenv()

# Import existing logic
import data_processor

# Page config
st.set_page_config(page_title="ASEAN Stock Analyzer", layout="wide")

# 環境変数からパスワードとAPIキーを取得
env_password = os.environ.get("APP_PASSWORD")
env_gemini_key = os.environ.get("GEMINI_API_KEY")

# --- 🔐 PASSWORD AUTHENTICATION ---
def password_entered():
    if st.session_state["password"] == env_password:
        st.session_state["password_correct"] = True
        del st.session_state["password"]
    else:
        st.session_state["password_correct"] = False

def check_password():
    if "password_correct" not in st.session_state:
        st.text_input("Please enter the password:", type="password", on_change=password_entered, key="password")
        return False
    elif not st.session_state["password_correct"]:
        st.text_input("Please enter the password:", type="password", on_change=password_entered, key="password")
        st.error("😕 Password incorrect")
        return False
    return True

if not check_password():
    st.stop()

# --- 💾 SESSION STATE INITIALIZATION ---
if "excel_buffer" not in st.session_state:
    st.session_state.excel_buffer = None
if "final_df" not in st.session_state:
    st.session_state.final_df = None

# --- 🛠 HELPERS ---
def clean_duplicate_columns(df, step_name=""):
    if df.columns.duplicated().any():
        duplicated_cols = df.columns[df.columns.duplicated()].tolist()
        if st.session_state.get("debug_mode"):
            st.warning(f"[{step_name}] 重複列を検出しました: {duplicated_cols}")
        return df.loc[:, ~df.columns.duplicated()].copy()
    return df

# --- MAIN APP ---
st.title("📊 ASEAN Stock Financial & AI Analysis Tool")

with st.sidebar:
    st.header("Settings")
    debug_mode = st.checkbox("Debug Mode (列名の状態を表示)", key="debug_mode")
    
    # st.secrets ではなく os.environ を使用するように修正
    if env_gemini_key:
        os.environ["GEMINI_API_KEY"] = env_gemini_key
        from google import genai
        data_processor.client = genai.Client(api_key=env_gemini_key)
        st.success("API Key loaded from .env ✅")
    else:
        api_key = st.text_input("Gemini API Key", type="password")
        if api_key:
            os.environ["GEMINI_API_KEY"] = api_key
            from google import genai
            data_processor.client = genai.Client(api_key=api_key)

uploaded_file = st.file_uploader("Upload Stock List (CSV)", type=["csv"])
use_sample = st.checkbox("Use default list (asean_list.csv) if no file is available")

# --- EXECUTE ANALYSIS ---
if st.button("Start Analysis 🚀"):
    target_csv = uploaded_file if uploaded_file else ("asean_list.csv" if use_sample else None)
    
    if target_csv is None:
        st.error("Please upload a CSV file.")
    else:
        try:
            st.session_state.excel_buffer = None
            df_input = pd.read_csv(target_csv, header=None)
            codes = df_input[0].astype(str).tolist()
            
            status_text = st.empty()
            progress_bar = st.progress(0)
            
            all_results = []
            for i, code in enumerate(codes):
                code = code.strip()
                status_text.text(f"Processing ({i+1}/{len(codes)}): {code}...")
                progress_bar.progress((i + 1) / (len(codes) + 1))
                raw_data = data_processor.get_stock_data(code)
                if raw_data:
                    all_results.append(data_processor.extract_data(code, raw_data))
                time.sleep(0.2)
            
            if all_results:
                status_text.text("🤖 Running AI Analysis...")
                all_results = data_processor.batch_analyze_segments(all_results)
                
                df = pd.DataFrame(all_results)
                df = clean_duplicate_columns(df, "DataFrame作成直後")
                
                status_text.text("📏 Formatting data...")
                df = data_processor.format_for_excel(df)
                df = clean_duplicate_columns(df, "format_for_excel後")
                
                df["Ref"] = range(1, len(df) + 1)
                empty_cols = ["Taka's comments", "Remarks", "Visited (V) / Meeting Proposal (MP)", "Access", "Last Communications", "Category Classification/\nShareInvestor", "Incorporated\n (IN / Year)", "Category Classification/SGX", "Sector & Industry/ SGX"]
                for col in empty_cols:
                    if col not in df.columns:
                        df[col] = ""
                df["Listed 'o' / Non Listed \"x\""] = "o"

                yesterday = datetime.now() - timedelta(days=1)
                yesterday_str = yesterday.strftime("%b %d")
                final_stock_price_col = f"Stock Price ({yesterday_str}, Closing)"
                final_rate_col = f"Exchange Rate (to SGD) ({yesterday_str}, Closing)"
                
                df = clean_duplicate_columns(df, "リネーム直前")
                rename_dict = {}
                if "Stock Price" in df.columns: rename_dict["Stock Price"] = final_stock_price_col
                if "Exchange Rate" in df.columns: rename_dict["Exchange Rate"] = final_rate_col
                if "Number of Employee" in df.columns: rename_dict["Number of Employee"] = "Number of Employee Current"
                df = df.rename(columns=rename_dict)
                
                df = clean_duplicate_columns(df, "リネーム直後")

                status_text.text("🔄 Reordering columns...")
                target_order = [
                    "Ref", "Name of Company", "Code", "Listed 'o' / Non Listed \"x\"", "Taka's comments", "Remarks", "Visited (V) / Meeting Proposal (MP)", "Website", "Major Shareholders", "Currency", final_rate_col, "FY", "REVENUE SGD('000)", "Segments", "PROFIT ('000)", "GROSS PROFIT ('000)", "OPERATING PROFIT ('000)", "NET PROFIT (Group) ('000)", "NET PROFIT (Shareholders) ('000)", "Minority Interest ('000)", "Shareholders' Equity ('000)", "Total Equity ('000)", "TOTAL ASSET ('000)", "Debt/Equity(%)", "Loan ('000)", "Loan/Equity (%)", final_stock_price_col, "Shares Outstanding ('000)", "Market Cap ('000)", "Summary of Business", "Chairman / CEO", "Address", "Contact No.", "Access", "Last Communications", "Number of Employee Current", "Category Classification/YahooFin", "Sector & Industry/YahooFin", "Category Classification/\nShareInvestor", "Incorporated\n (IN / Year)", "Category Classification/SGX", "Sector & Industry/ SGX"
                ]
                
                for col in target_order:
                    if col not in df.columns:
                        df[col] = ""
                
                df = clean_duplicate_columns(df, "Reindex直前最終チェック")
                
                if debug_mode:
                    st.write("Current Columns:", df.columns.tolist())

                df = df.reindex(columns=target_order)

                status_text.text("💾 Generating Excel file...")
                temp_buffer = io.BytesIO()
                df.to_excel(temp_buffer, index=False)
                temp_buffer.seek(0)
                
                wb = load_workbook(temp_buffer)
                ws = wb.active
                header_fill = PatternFill(start_color="fefe99", end_color="fefe99", fill_type="solid")
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = Font(bold=True)
                
                final_buffer = io.BytesIO()
                wb.save(final_buffer)
                
                st.session_state.excel_buffer = final_buffer.getvalue()
                st.session_state.final_df = df
                st.session_state.output_filename = f"asean_financial_data_{datetime.today().strftime('%Y-%m-%d')}.xlsx"
                
                progress_bar.progress(100)
                status_text.text("✅ All processes completed!")

        except Exception as e:
            st.error(f"❌ Error during processing: {e}")
            if debug_mode:
                st.exception(e)

# --- 📥 DOWNLOAD AREA ---
if st.session_state.excel_buffer is not None:
    st.divider()
    st.success("Analysis results ready!")
    
    st.download_button(
        label="📥 Download Excel File",
        data=st.session_state.excel_buffer,
        file_name=st.session_state.output_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_btn"
    )
    
    st.subheader("Data Preview")
    st.dataframe(st.session_state.final_df)