# main_sector.py
import pandas as pd
import datetime
from pathlib import Path
import sys
import time
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font

import yfinance_client
import data_processor
import asean_stock_codes 

def main():
    print("=== 国・セクター別 ASEAN株 財務データ取得システム (AIセグメント分析対応版) ===")
    
    # ---------------------------------------------------------
    # 1. 国（取引所）の選択
    # ---------------------------------------------------------
    print("\n[ステップ1] 対象の国（取引所）を選択してください。")
    print("利用可能な国コード:")
    print("  SG: シンガポール (.SI)")
    print("  MY: マレーシア (.KL)")
    print("  ID: インドネシア (.JK)")
    print("  TH: タイ (.BK)")
    print("  PH: フィリピン (.PS)")
    print("  VN: ベトナム (.VN)")
    
    country_input = input("国コードを入力してください (カンマ区切りで複数可, 例: SG, MY): ").strip().upper()
    
    if not country_input:
        print("国コードが入力されませんでした。終了します。")
        return

    target_countries = [c.strip() for c in country_input.split(',') if c.strip()]
    
    suffix_map = {
        "SG": ".SI", "MY": ".KL", "ID": ".JK", 
        "TH": ".BK", "VN": ".VN", "PH": ".PS"
    }
    
    target_suffixes = []
    for country in target_countries:
        if country in suffix_map:
            target_suffixes.append(suffix_map[country])
        else:
            print(f"警告: 未対応またはリストにない国コード '{country}' は無視されます。")
    
    if not target_suffixes:
        print("有効な国コードが選択されませんでした。")
        return

    # ---------------------------------------------------------
    # 2. ターゲットとするセクターを指定
    # ---------------------------------------------------------
    print("\n[ステップ2] 取得したいセクター名を入力してください。")
    print("例: Technology, Real Estate, Financial Services")
    sector_input = input("セクター名 (カンマ区切りで複数可): ").strip()
    
    if not sector_input:
        print("セクター名が入力されませんでした。終了します。")
        return

    target_sectors = [s.strip() for s in sector_input.split(',') if s.strip()]
    print(f"\nターゲットセクター: {target_sectors}")

    # ---------------------------------------------------------
    # 3. リストからのフィルタリング
    # ---------------------------------------------------------
    print("内蔵リストから対象国の銘柄を抽出しています...")
    
    all_codes = asean_stock_codes.ALL_ASEAN_CODES
    country_filtered_codes = []
    
    for code in all_codes:
        if any(code.endswith(suffix) for suffix in target_suffixes):
            country_filtered_codes.append(code)
            
    print(f"国フィルター適用後: {len(country_filtered_codes)} 件の銘柄が対象です。")
    
    if not country_filtered_codes:
        print("対象となる銘柄が見つかりませんでした。")
        return

    # ---------------------------------------------------------
    # 4. セクターによるスクリーニング
    # ---------------------------------------------------------
    target_codes = []
    import yfinance as yf

    print("セクター検索を開始します (yfinance)...")
    for i, code in enumerate(country_filtered_codes):
        print(f"\rスクリーニング中: {i+1}/{len(country_filtered_codes)} ({code})", end="")
        
        try:
            ticker = yf.Ticker(code)
            info = ticker.info
            company_sector = info.get('sector', 'Unknown')
            
            is_match = False
            for target in target_sectors:
                if target.lower() in str(company_sector).lower():
                    is_match = True
                    break
            
            if is_match:
                target_codes.append(code)
                print(f"\n  -> Hit! {code}: {info.get('longName')} ({company_sector})")
                
        except Exception:
            pass 
            
        time.sleep(0.1)

    print(f"\n\n検索終了。該当銘柄数: {len(target_codes)} 件")
    
    if len(target_codes) == 0:
        print("該当する銘柄が見つかりませんでした。")
        return

    # ---------------------------------------------------------
    # 5. 詳細データ取得
    # ---------------------------------------------------------
    print("\n詳細データの取得を開始します...")
    
    all_results = []
    
    for code in target_codes:
        print(f"\n--- {code} の処理中 ---")
        
        raw_data = yfinance_client.get_stock_data(code)
        
        if raw_data:
            processed_data = data_processor.extract_data(code, raw_data)
            all_results.append(processed_data)
            print(f"  会社名: {processed_data.get('Name of Company')}")
            print(f"  売上高: {processed_data.get('REVENUE')}")
        else:
            print("  データの取得に失敗しました。")
            
        time.sleep(0.5)

    # ---------------------------------------------------------
    # 6. AIによるセグメント分析
    # ---------------------------------------------------------
    if all_results:
        print("\n--- 全データ取得完了。AIによるセグメント分析を開始します ---")
        all_results = data_processor.batch_analyze_segments(all_results)

    # ---------------------------------------------------------
    # 7. Excel生成
    # ---------------------------------------------------------
    if all_results:
        print("\nExcelファイルを作成しています...")
        df = pd.DataFrame(all_results)
        
        # 1. 数値整形
        df = data_processor.format_for_excel(df)
        
        # 2. 不要列削除
        if "Sector /Industry" in df.columns:
            df = df.drop(columns=["Sector /Industry"])
            
        # 3. Ref列追加など
        df["Ref"] = range(1, len(df) + 1)

        empty_cols = [
            "Taka's comments", "Remarks", "Visited (V) / Meeting Proposal (MP)", 
            "Access", "Last Communications", "Category Classification/\nShareInvestor", 
            "Incorporated\n (IN / Year)", "Category Classification/SGX", "Sector & Industry/ SGX"
        ]
        
        for col in empty_cols:
            df[col] = ""
            
        df["Listed 'o' / Non Listed \"x\""] = "o"

        # ★追加: Stock Priceのカラム名を動的に特定する
        # "Stock Price" で始まるカラムを探す
        stock_price_col = next((col for col in df.columns if str(col).startswith("Stock Price")), "Stock Price")

        target_order = [
            "Ref",
            "Name of Company",
            "Code",
            "Listed 'o' / Non Listed \"x\"",
            "Taka's comments",
            "Remarks",
            "Visited (V) / Meeting Proposal (MP)",
            "Website",
            "Major Shareholders",
            "Currency",
            "Exchange Rate (to SGD)",
            "FY",
            "REVENUE SGD('000)", 
            "Segments", 
            "PROFIT ('000)", 
            "GROSS PROFIT ('000)", 
            "OPERATING PROFIT ('000)", 
            "NET PROFIT (Group) ('000)", 
            "NET PROFIT (Shareholders) ('000)", 
            "Minority Interest ('000)", 
            "Shareholders' Equity ('000)", 
            "Total Equity ('000)", 
            "TOTAL ASSET ('000)", 
            "Debt/Equity(%)", 
            "Loan ('000)", 
            "Loan/Equity (%)",
            
            # ★★★ 追加した3項目 ★★★
            stock_price_col,              # 株価 (日付入り)
            "Shares Outstanding ('000)",  # 発行株式数
            "Market Cap ('000)",          # 時価総額
            
            "Summary of Business", 
            "Chairman / CEO", 
            "Address", 
            "Contact No.", 
            "Access", 
            "Last Communications", 
            "Number of Employee", 
            "Category Classification/YahooFin", 
            "Sector & Industry/YahooFin", 
            "Category Classification/\nShareInvestor", 
            "Incorporated\n (IN / Year)", 
            "Category Classification/SGX", 
            "Sector & Industry/ SGX"
        ]
        
        # カラムが存在しない場合の補完
        for col in target_order:
             if col not in df.columns:
                 df[col] = ""
        
        df = df.reindex(columns=target_order)
        df = df.rename(columns={"Number of Employee": "Number of Employee Current"})

        # ファイル名生成
        countries_str = "_".join(target_countries)
        if len(target_sectors) == 1:
            sector_name_for_file = target_sectors[0].replace(" ", "_").replace("/", "-")
        else:
            sector_name_for_file = "Multi_Sectors"

        today = datetime.date.today().strftime("%Y-%m-%d")
        base_name = f"asean_data_{countries_str}_{sector_name_for_file}_{today}"
        filename = f"{base_name}.xlsx"
        
        counter = 1
        while Path(filename).exists():
            filename = f"{base_name}_{counter}.xlsx"
            counter += 1
            
        try:
            # 1. 保存
            df.to_excel(filename, index=False)
            
            # 2. 書式設定
            wb = load_workbook(filename)
            ws = wb.active
            right_align = Alignment(horizontal='right')
            
            header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            header_font = Font(bold=True)

            for cell in ws[1]:
                col_name = str(cell.value)
                col_idx = cell.column
                
                cell.fill = header_fill
                cell.font = header_font
                
                number_format = None
                apply_alignment = False
                
                if "('000)" in col_name:
                    number_format = '#,##0;(#,##0)'
                    apply_alignment = True
                elif "(%)" in col_name or "%" in col_name:
                    number_format = '0.00%'
                    apply_alignment = True
                elif col_name == "FY":
                    apply_alignment = True
                # ★追加: Stock Priceも数値フォーマット
                elif "Stock Price" in col_name:
                    number_format = '#,##0.000' # 小数点3桁まで表示
                    apply_alignment = True
                
                if number_format or apply_alignment:
                    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                        for cell_data in row:
                            if apply_alignment:
                                cell_data.alignment = right_align
                            if number_format:
                                cell_data.number_format = number_format
            
            wb.save(filename)
            print(f"★★★ 成功: {filename} に保存しました ★★★")
            
        except Exception as e:
            print(f"エラー: Excel保存に失敗しました ({e})")
            
    else:
        print("保存するデータがありませんでした。")

if __name__ == "__main__":
    main()